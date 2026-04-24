"""
================================================================================
VALIDAÇÃO DE EXTRATOS BANCÁRIOS v8.0
Stoic Capital
================================================================================

ARQUITETURA DO PIPELINE
-----------------------
O sistema segue o padrão Collect → Normalize → Classify → Validate → Deliver:

  1. Upload       Operador envia arquivos CSV/XLSX de diferentes bancos e unidades.
  2. Verificação  Sistema detecta a unidade de negócio por fuzzy matching no nome
                  do arquivo. Operador confirma ou corrige.
  3. Processamento Para cada linha de cada arquivo:
                  (a) Detecta o schema (mapeamento coluna → anchor padronizado)
                  (b) Extrai os campos âncora (Data, Valor, Tipo, Descricao, ...)
                  (c) Aplica fallbacks para padrões conhecidos de bancos específicos
                  (d) Classifica o lançamento em Grupo A (confiável) ou Grupo B
                      (requer revisão humana)
  4. Revisão      Operador valida o Grupo B: confirma, edita ou exclui.
                  Ações em lote disponíveis por tipo de problema, arquivo ou unidade.
  5. Exportação   Excel com 4 abas: Legenda, BD_Extratos, De_Para, Sumario.

SCHEMA ÂNCORA (colunas padronizadas de saída)
----------------------------------------------
  Obrigatórios: Data, Valor, Descricao
  Calculados:   Tipo (entrada | saida | indefinido)
  Opcionais:    Conta, Banco, CNPJ, Centro_Custo

  Qualquer lançamento com campo obrigatório ausente vai para Grupo B.

PERSISTÊNCIA
------------
  Sessão (em memória, por restart):  arquivos, lançamentos, schema_map
  Config (config.json, permanente):  unidades, depara, fn_patterns

FALLBACKS DE BANCOS ESPECÍFICOS (ver loop em processar())
---------------------------------------------------------
  B1 — BRB XLS:           Linhas de cabeçalho do banco são descartadas automaticamente.
  B2 — Stone Comprovante: Valor calculado de |Saldo Depois − Saldo Antes|;
                          Tipo inferido de extra_Movimentação.
  B3 — Cora / GN:         Descrição montada de extra_Transação + extra_Identificação.

DEPENDÊNCIAS
------------
  flask, pandas, openpyxl, xlrd, werkzeug  (requirements.txt)
================================================================================
"""

import os, json, re, difflib, hashlib, math
from pathlib import Path
from flask import Flask, request, jsonify, render_template, send_file
import pandas as pd
from werkzeug.utils import secure_filename
from datetime import datetime
import io
from ofxparse import OfxParser

from detectores_banco import detectar_cabecalho
from resolucao_conta  import resolver_conta, validar_cruzado

# ============================================================
# CONFIGURAÇÃO FLASK
# ============================================================

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # limite de upload: 100 MB
UPLOAD_FOLDER = Path(__file__).parent / 'uploads'
UPLOAD_FOLDER.mkdir(exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
CONFIG_PATH = Path(__file__).parent / 'config.json'
ALLOWED_EXT = {'csv', 'xls', 'xlsx', 'pdf', 'ofx'}

# Importação tolerante do extrator de PDF (módulo opcional)
try:
    from pdf_extractor import extrair_pdf as _extrair_pdf
    HAS_PDF_EXTRACTOR = True
except Exception as _e:
    HAS_PDF_EXTRACTOR = False
    print(f"[PDF] Extrator indisponível: {_e}")

# ── Encoder defensivo contra float NaN ───────────────────────────────
# O pandas mantém float('nan') em células vazias mesmo com dtype=str.
# json.dumps converte NaN para o token JS `NaN`, que não é JSON válido,
# causando falha silenciosa no browser (JSON.parse rejeita o response inteiro).
# Solução: substituir NaN/Inf por None (→ null) antes de qualquer serialização.

def _nan_clean(obj):
    """Percorre recursivamente um objeto e substitui float NaN/Inf por None."""
    if isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
        return None
    if isinstance(obj, dict):
        return {k: _nan_clean(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_nan_clean(v) for v in obj]
    return obj

# Sobrescreve jsonify para sempre passar pelo _nan_clean
_orig_jsonify = jsonify
def jsonify(*args, **kwargs):
    if len(args) == 1 and not kwargs:
        data = _nan_clean(args[0])
    else:
        data = _nan_clean(dict(*args, **kwargs))
    from flask import current_app
    resp = current_app.response_class(
        json.dumps(data, ensure_ascii=False),
        mimetype='application/json'
    )
    return resp


# ============================================================
# SCHEMA ÂNCORA — mapeamento de sinônimos por banco/sistema
# ============================================================
# Cada chave é o nome do campo padronizado no output (anchor).
# Cada valor é a lista de variações de nomes que diferentes bancos/sistemas usam.
# detectar_mapa() usa esta tabela para mapear colunas dos arquivos → anchors.

ANCHOR_MAPS = {
    'Data': [
        'data', 'dt', 'date', 'data_lancamento', 'data_pagamento', 'dt_lancamento',
        'dt_pagamento', 'data_mov', 'data_transacao', 'data_extrato', 'data_operacao',
        'datalancamento', 'datapagamento', 'datamov', 'dataoperacao', 'data_documento',
    ],
    'Valor': [
        'valor', 'value', 'amount', 'vlr', 'vlr_lancamento', 'vl_lancamento',
        'importancia', 'montante', 'valorlancamento', 'valortransacao', 'vlrlancamento',
        'valor_transacao', 'valor_operacao',
    ],
    # Débito/Crédito separados — padrão de alguns bancos (ex: BRB, BB)
    'Debito': [
        'debito', 'saida', 'debit', 'saidas', 'valor_debito', 'vlr_debito',
        'debitos', 'debitobrl', 'vlrdebito', 'saidas_r',
    ],
    'Credito': [
        'credito', 'entrada', 'credit', 'entradas', 'valor_credito', 'vlr_credito',
        'creditos', 'creditobrl', 'vlrcredito', 'entradas_r',
    ],
    # Tipo textual do lançamento (D/C, DEBIT/CREDIT, etc.)
    'Tipo': [
        'tipo', 'type', 'natureza', 'trntype', 'dc', 'debito_credito',
        'tipo_lancamento', 'tipo_transacao', 'natureza_lancamento', 'indicador',
        'movimentacao',
    ],
    'Descricao': [
        'descricao', 'description', 'memo', 'historico', 'complemento',
        'detalhe', 'detalhes', 'nome', 'descr', 'discriminacao', 'lancamento',
        'historico_lancamento', 'descricao_lancamento', 'detalhe_lancamento',
        'historico_extrato', 'historico_banco',
    ],
    'Conta': [
        'conta', 'account', 'nr_conta', 'numero_conta', 'conta_corrente',
        'agencia_conta', 'nrconta', 'numeroconta', 'conta_banco',
    ],
    'Banco': [
        'banco', 'bank', 'instituicao', 'origem', 'banco_origem', 'nome_banco',
        'instituicao_financeira', 'instituicaofinanceira',
    ],
    'CNPJ': [
        'cnpj', 'cpf_cnpj', 'documento', 'cpf', 'cnpj_cpf', 'cpfcnpj',
    ],
    'Centro_Custo': [
        'centro_custo', 'cost_center', 'cc', 'centro', 'ccusto', 'centrocusto',
        'centro_de_custo', 'cod_cc',
    ],
}

# Anchors que DEVEM estar presentes para o lançamento ser Grupo A
ANCHORS_OBRIGATORIOS = {'Data', 'Valor', 'Descricao'}
# Anchors opcionais — ausência não penaliza
ANCHORS_OPCIONAIS    = {'Tipo', 'Conta', 'Banco', 'CNPJ', 'Centro_Custo'}


# ============================================================
# UNIDADES PADRÃO
# ============================================================

DEFAULT_UNIDADES = [
    {'id': 'GN',    'marca': 'GN',         'desc_unidade': 'GN CSC'},
    {'id': 'LD1',   'marca': 'Ludika',      'desc_unidade': 'Ludika Asa-Sul'},
    {'id': 'LD2',   'marca': 'Ludika',      'desc_unidade': 'Ludika [Outra]'},
    {'id': 'MN303', 'marca': 'Manta',       'desc_unidade': 'Manta 303'},
    {'id': 'MN610', 'marca': 'Manta',       'desc_unidade': 'Manta 610'},
    {'id': 'MN702', 'marca': 'Manta',       'desc_unidade': 'Manta 702'},
    {'id': 'MN712', 'marca': 'Manta',       'desc_unidade': 'Manta 712'},
    {'id': 'MNJB',  'marca': 'Manta',       'desc_unidade': 'Manta JB'},
    {'id': 'MNVP',  'marca': 'Manta',       'desc_unidade': 'Manta VP'},
    {'id': 'RC608', 'marca': 'RaiaClube',   'desc_unidade': 'RaiaClube 608'},
    {'id': 'RCA',   'marca': 'RaiaClube',   'desc_unidade': 'RaiaClube ASTCU'},
]


# ============================================================
# PERSISTÊNCIA — config.json
# ============================================================

def load_config() -> dict:
    """
    Carrega configurações do config.json.
    Retorna defaults se o arquivo não existir ou estiver corrompido.
    Garante backward compatibility adicionando chaves ausentes.

    Migração v9: 'contas_bancarias' é criada como [] se ausente.
    fn_patterns é mantida intacta por compat com fluxo v8 em modo_legado,
    mas registra aviso em NOTAS_REFATORACAO.md na primeira carga onde
    contas_bancarias está vazia e fn_patterns tem entradas.
    """
    if CONFIG_PATH.exists():
        try:
            c = json.loads(CONFIG_PATH.read_text('utf-8'))
            if 'unidades'         not in c: c['unidades']         = DEFAULT_UNIDADES
            if 'depara'           not in c: c['depara']           = {}
            if 'fn_patterns'      not in c: c['fn_patterns']      = {}
            if 'contas_bancarias' not in c:
                c['contas_bancarias'] = []
                _aviso_migracao_v9(c)
            return c
        except Exception as e:
            print(f"[CONFIG] Erro ao carregar: {e}")
    return {'unidades': DEFAULT_UNIDADES, 'depara': {}, 'fn_patterns': {},
            'contas_bancarias': []}


def _aviso_migracao_v9(cfg: dict) -> None:
    """Anexa aviso em NOTAS_REFATORACAO.md quando config v8 é lida.

    Chamado uma única vez no upgrade. Silencioso se o arquivo de notas
    não existir (ambiente não-dev).
    """
    if not cfg.get('fn_patterns'):
        return
    notas = Path(__file__).parent / 'NOTAS_REFATORACAO.md'
    if not notas.exists():
        return
    stamp = datetime.now().strftime('%Y-%m-%d')
    marcador = '## Migração v8 → v9 detectada'
    try:
        conteudo_atual = notas.read_text('utf-8')
        if marcador in conteudo_atual:
            return
        bloco = (
            f"\n\n---\n\n{marcador} ({stamp})\n\n"
            f"- `config.json` foi carregada pela primeira vez com "
            f"`contas_bancarias` vazia e `fn_patterns` populada ({len(cfg['fn_patterns'])} entradas).\n"
            f"- `fn_patterns` são legado v8 e devem ser substituídos por "
            f"`contas_bancarias`. Enquanto não houver contas cadastradas, "
            f"o sistema opera em **modo_legado** (Rastreab=BAIXA em tudo).\n"
            f"- Entradas atuais de fn_patterns (para migração manual):\n"
            + '\n'.join(f"  - `{k}` → unit_id `{v}`"
                        for k, v in cfg['fn_patterns'].items())
            + "\n"
        )
        notas.write_text(conteudo_atual + bloco, 'utf-8')
    except Exception as e:
        print(f"[MIGRACAO_V9] Aviso não pôde ser registrado: {e}")

def save_config(cfg: dict) -> None:
    """Persiste configuração em config.json (UTF-8, indentado)."""
    try:
        CONFIG_PATH.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), 'utf-8')
    except Exception as e:
        print(f"[CONFIG] Erro ao salvar: {e}")

CFG = load_config()


# ============================================================
# SESSÃO EM MEMÓRIA
# ============================================================
# SESSION é reiniciado a cada restart do Flask.
# Para persistência multi-sessão, migrar para SQLite.

SESSION = {
    'arquivos':        [],    # [{filename, size, path, hash}]
    'lancamentos':     [],    # [{id, arquivo, Data, Valor, Tipo, ...}]
    'schema_map':      {},    # {col_original: anchor} da última execução
    'processado':      False,
    'doc_verificados': {},    # {filename: unit_id} — confirmações manuais (legado v8)
    'previews':        {},    # {filename: dict} — cache de sumários
    'progresso':       {'pct': 0, 'msg': '', 'ativo': False},

    # v9: identificação e controle de conta bancária por arquivo.
    'deteccao_cab':         {},  # {filename: dict do detectar_cabecalho}
    'cruzamento':           {},  # {filename: dict do validar_cruzado}
    'conta_por_arquivo':    {},  # {filename: conta_id} — resoluções manuais
    'conta_por_arquivo_meta': {},  # {filename: {metodo_original, escolha_humana, motivo, timestamp}}
    'arquivos_bloqueados':  [],  # lista de filenames bloqueados no último processar()
    'audit_log':            [],  # [{ts, filename, acao, detalhes}]
}


# ============================================================
# HELPERS — NORMALIZAÇÃO
# ============================================================

def norm_col(col: str) -> str:
    """
    Normaliza nome de coluna: minúsculo, sem acentos, sem pontuação.
    Ex: 'Valor Lançamento' → 'valorlancamento'
    """
    s = str(col).lower().strip()
    for a, b in [('á','a'),('â','a'),('ã','a'),('à','a'),('é','e'),('ê','e'),
                 ('í','i'),('ó','o'),('ô','o'),('õ','o'),('ú','u'),('ç','c')]:
        s = s.replace(a, b)
    return re.sub(r'[^a-z0-9]', '', s)

def detectar_mapa(colunas: list) -> dict:
    """
    Mapeia colunas de um arquivo para anchors padronizados.

    Três camadas:
      1. De-para salvo em config.json (aprendido de execuções anteriores)
      2. Match exato em ANCHOR_MAPS (após normalização)
      3. Fuzzy matching (SequenceMatcher, limiar 0.82)

    Returns:
        {col_original: anchor | None}  — None = coluna não reconhecida
    """
    depara    = CFG.get('depara', {})
    mapa      = {}
    syn_index = {}
    for anchor, syns in ANCHOR_MAPS.items():
        for s in syns:
            syn_index[norm_col(s)] = anchor

    for col in colunas:
        cn = norm_col(col)
        if col in depara:
            mapa[col] = depara[col]; continue
        if cn in syn_index:
            mapa[col] = syn_index[cn]; continue
        best, bscore = None, 0
        for nsyn, anchor in syn_index.items():
            sc = difflib.SequenceMatcher(None, cn, nsyn).ratio()
            if sc > bscore:
                bscore = sc; best = anchor
        mapa[col] = best if bscore >= 0.82 else None
    return mapa

def inv_mapa(schema_map: dict) -> dict:
    """
    Inverte schema_map para lookup direto por anchor.
    Ex: {'Vlr': 'Valor', 'DT': 'Data'} → {'Valor': 'Vlr', 'Data': 'DT'}
    Primeira ocorrência de cada anchor tem prioridade.
    """
    inv = {}
    for col, anchor in schema_map.items():
        if anchor and anchor not in inv:
            inv[anchor] = col
    return inv


# ============================================================
# HELPERS — PARSING
# ============================================================

def parse_valor(v) -> float | None:
    """
    Converte valor monetário (string ou número) para float.

    Regra de separador: o ÚLTIMO símbolo ',' ou '.' na string é o
    separador decimal; os anteriores (do mesmo tipo ou diferentes) são
    separadores de milhar. Aceita formato BR ('1.500,00'), US/canônico
    ('1,500.00') e variações misturadas ('1.500.000,00', '1,500,000.00').

    Sinais aceitos:
      - prefixo '+' ou '-':  "-1.500,00"
      - sufixo '-' (BRB):    "84.000,00-"
      - parênteses contábeis: "(1.500,00)" → -1500.0

    Retorna None para NaN, vazio, None ou string não-parseável.
    """
    if v is None:
        return None
    s = str(v).strip()
    if not s or s.lower() in ('nan', 'nat', 'none'):
        return None

    # Convenção contábil: (1.500,00) → -1500.00
    sinal = 1
    if s.startswith('(') and s.endswith(')'):
        sinal = -1
        s = s[1:-1].strip()
    s = re.sub(r'[R$\s\u00a0\xa0]', '', s)

    # Sinal negativo posterior (BRB): "84.000,00-"
    if s.endswith('-'):
        sinal = -sinal
        s = s.rstrip('-')

    # Sinal explícito no início
    if s.startswith('-'):
        sinal = -sinal
        s = s[1:]
    elif s.startswith('+'):
        s = s[1:]

    if not s:
        return None

    # Último ',' ou '.' é decimal; os anteriores são milhar.
    last_comma = s.rfind(',')
    last_dot   = s.rfind('.')
    if last_comma > last_dot:
        s = s.replace('.', '').replace(',', '.')
    elif last_dot > last_comma:
        s = s.replace(',', '')
    # else: nenhum separador — inteiro puro

    try:
        return float(s) * sinal
    except ValueError:
        return None

def parse_data(v, preservar_raw: bool = False):
    """
    Converte data em múltiplos formatos para ISO YYYY-MM-DD.

    Retorna None quando:
      - entrada é vazia/None/NaN/NaT
      - nenhum formato reconhecido funciona
      - data é calendaricamente inválida (ex: '00/00/0000', '31/02/2026')

    Se preservar_raw=True, retorna tupla (iso_ou_None, string_original).
    Isso permite distinguir 'data_ausente' (raw vazio) de 'data_invalida'
    (raw preenchido mas não parseável) no classificador.
    """
    raw = '' if v is None else str(v).strip()

    def _ret(valid):
        return (valid, raw) if preservar_raw else valid

    if v is None:
        return _ret(None)

    if isinstance(v, (pd.Timestamp, datetime)):
        try:
            if pd.isna(v):
                return _ret(None)
            return _ret(v.strftime('%Y-%m-%d'))
        except (ValueError, AttributeError):
            return _ret(None)

    s = raw
    if not s or s.lower() in ('nan', 'nat', 'none', 'nd'):
        return _ret(None)

    for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d.%m.%Y',
                '%Y%m%d', '%d/%m/%y', '%m/%d/%Y']:
        try:
            return _ret(datetime.strptime(s[:10], fmt).strftime('%Y-%m-%d'))
        except ValueError:
            pass
    return _ret(None)


# ============================================================
# HELPERS — DETECÇÃO DE TIPO
# ============================================================

def detectar_tipo(row: dict, inv: dict) -> tuple[str, str]:
    """
    Determina direção do lançamento por cascata de 4 estratégias:

      1. Colunas separadas Débito/Crédito (alta confiança)
      2. Coluna textual de tipo D/C (alta confiança)
      3. Sinal do valor — positivo=entrada, negativo=saída (média)
      4. Indefinido → issue 'tipo_indefinido', lançamento vai para B

    Returns: (tipo, confianca) onde tipo ∈ {'entrada','saida','indefinido'}
    """
    col_deb, col_cred = inv.get('Debito'), inv.get('Credito')
    if col_deb and col_cred:
        d  = parse_valor(row.get(col_deb))
        c  = parse_valor(row.get(col_cred))
        dv = abs(d) if d is not None else 0
        cv = abs(c) if c is not None else 0
        if dv > 0 and cv == 0: return 'saida',  'alta'
        if cv > 0 and dv == 0: return 'entrada', 'alta'
        if dv > 0: return 'saida',  'alta'
        if cv > 0: return 'entrada', 'alta'

    col_tipo = inv.get('Tipo')
    if col_tipo:
        v = str(row.get(col_tipo, '')).strip().upper()
        MAPA_TIPO = {
            'D':'saida','DEB':'saida','DEBITO':'saida','DÉBITO':'saida','DEBIT':'saida',
            'SAIDA':'saida','SAÍDA':'saida','-1':'saida',
            'C':'entrada','CRED':'entrada','CREDITO':'entrada','CRÉDITO':'entrada','CREDIT':'entrada',
            'ENTRADA':'entrada','+1':'entrada','DEPÓSITO':'entrada','DEPOSITO':'entrada',
        }
        if v in MAPA_TIPO:
            return MAPA_TIPO[v], 'alta'

    # Infere tipo pela descrição (BRB PDFs onde sinal fica em coluna separada)
    col_desc = inv.get('Descricao')
    if col_desc:
        desc_up = str(row.get(col_desc, '')).strip().upper()
        if desc_up.startswith('DEBITO') or desc_up.startswith('DÉBITO'):
            return 'saida', 'media'
        if desc_up.startswith('CREDITO') or desc_up.startswith('CRÉDITO'):
            return 'entrada', 'media'

    col_val = inv.get('Valor')
    if col_val:
        n = parse_valor(row.get(col_val))
        if n is not None:
            return ('saida' if n < 0 else 'entrada'), 'media'

    return 'indefinido', 'baixa'


# ============================================================
# HELPERS — CLASSIFICAÇÃO A / B
# ============================================================

def classificar(row_norm: dict, unit_conf: int, unit_confirmada: bool) -> tuple[str, str, list]:
    """
    Classifica lançamento em Grupo A (confiável) ou B (revisão humana).

    Critérios para Grupo B (qualquer um):
      - data_ausente:      Data None E Data_Raw vazio (campo fonte vazio)
      - data_invalida:     Data None E Data_Raw preenchido (parse falhou)
      - valor_ausente:     Valor None após todos os fallbacks
      - descricao_ausente: Descrição vazia após todos os fallbacks
      - tipo_indefinido:   Nenhuma estratégia de tipo funcionou
      - unidade_incerta:   Não confirmada manualmente E fuzzy < 80%

    data_ausente e data_invalida são mutuamente exclusivos.

    Confiabilidade:
      ALTA  → Grupo A
      MÉDIA → Grupo B com apenas tipo_indefinido
      BAIXA → Grupo B com campo obrigatório ausente (inclui data_invalida)

    Returns: (grupo, confiabilidade, issues)
    """
    issues = []
    if row_norm.get('Data') is None:
        data_raw = str(row_norm.get('Data_Raw', '') or '').strip()
        issues.append('data_invalida' if data_raw else 'data_ausente')
    if row_norm.get('Valor') is None:
        issues.append('valor_ausente')
    if not str(row_norm.get('Descricao', '')).strip():
        issues.append('descricao_ausente')
    if row_norm.get('Tipo') == 'indefinido':
        issues.append('tipo_indefinido')
    if not unit_confirmada and unit_conf < 80:
        issues.append('unidade_incerta')

    grupo = 'A' if not issues else 'B'
    if grupo == 'A':
        conf = 'ALTA'
    elif issues == ['tipo_indefinido']:
        conf = 'MÉDIA'
    else:
        conf = 'BAIXA'
    return grupo, conf, issues


# ============================================================
# HELPERS — LEITURA DE ARQUIVOS
# ============================================================

_HEADER_KEYS = re.compile(
    r'data|hist[óo]rico|descri[çc][ãa]o|valor|saldo|d[ée]bito|cr[ée]dito|'
    r'lan[çc]amento|movimenta[çc][ãa]o|transa[çc][ãa]o|documento',
    re.IGNORECASE,
)

_DATE_RX = re.compile(r'\d{1,2}[/\-.]\d{1,2}[/\-.]\d{2,4}|\d{4}-\d{2}-\d{2}')

def _split_merged_header(raw_row: list, ncols: int) -> list:
    """Quando a linha de header é uma única célula merged ('Data  Histórico  Valor'),
    quebra em pedaços por 2+ espaços e distribui pelas n colunas reais."""
    texto = ' '.join(str(c).strip() for c in raw_row if pd.notna(c) and str(c).strip())
    if not texto:
        return [f'col_{i}' for i in range(ncols)]
    pedacos = [p for p in re.split(r'\s{2,}|\t|\n', texto) if p.strip()]
    if len(pedacos) >= 2:
        if len(pedacos) < ncols:
            pedacos += [f'col_{i}' for i in range(len(pedacos), ncols)]
        return pedacos[:ncols]
    return [f'col_{i}' for i in range(ncols)]

def _read_excel_smart_header(filepath: str) -> pd.DataFrame:
    """Detecta a linha de header em xlsx/xls. Para BRB e Santander pdf-gerado
    o header não está na linha 0. Estratégia:
      1. Encontrar primeira linha com data → essa é a linha de dados.
      2. Header = linha imediatamente anterior (com mais keywords).
      3. Se header tiver poucas células (merged), quebrar por espaços múltiplos.
    """
    raw = pd.read_excel(filepath, dtype=str, header=None)
    if raw.empty:
        return pd.DataFrame()

    # Identifica linhas com data nas primeiras células (transação, não metadado).
    def _row_tem_data(i):
        for c in raw.iloc[i].tolist()[:4]:
            if pd.notna(c):
                s = str(c).strip()
                if _DATE_RX.fullmatch(s):
                    return True
                p = s.split()[0] if s else ''
                if p and _DATE_RX.fullmatch(p):
                    return True
        return False

    # Início real dos dados = primeira posição com >=3 linhas de data consecutivas
    # (ou >=2 se o arquivo for muito pequeno). Evita falso positivo em linhas
    # de metadado tipo "Data de Emissão: 2025-01-31".
    def _row_vazia(i):
        return all(pd.isna(c) or not str(c).strip() for c in raw.iloc[i].tolist())

    primeira_data = None
    limite = min(60, len(raw))
    min_consec = 2
    for i in range(limite):
        if _row_tem_data(i):
            # Conta linhas-data dentro de janela curta, tolerando linhas vazias
            consec = 1
            for j in range(i + 1, min(i + 6, len(raw))):
                if _row_tem_data(j):
                    consec += 1
                elif _row_vazia(j):
                    continue
                else:
                    break
            if consec >= min_consec:
                primeira_data = i
                break

    if primeira_data is None or primeira_data == 0:
        return pd.read_excel(filepath, dtype=str).fillna('')

    # Procura header nas linhas anteriores (até 5 acima): a melhor por keywords
    melhor_idx, melhor_score = primeira_data - 1, -1
    for i in range(max(0, primeira_data - 5), primeira_data):
        celulas = [str(c).strip() for c in raw.iloc[i].tolist() if pd.notna(c) and str(c).strip()]
        if not celulas:
            continue
        texto = ' '.join(celulas)
        score = len(_HEADER_KEYS.findall(texto))
        if score > melhor_score:
            melhor_score, melhor_idx = score, i

    # Mantém só colunas com dados na faixa de transações (descarta colunas
    # totalmente vazias intercaladas, comum em BRB Ludika).
    data_rows = raw.iloc[primeira_data:].copy()
    cols_uteis_idx = [j for j in range(data_rows.shape[1])
                      if data_rows.iloc[:, j].notna().any()
                      and (data_rows.iloc[:, j].astype(str).str.strip() != '').any()]
    data_rows = data_rows.iloc[:, cols_uteis_idx]
    ncols = data_rows.shape[1]

    raw_header_full = raw.iloc[melhor_idx].tolist()
    raw_header = [raw_header_full[j] for j in cols_uteis_idx]
    cells_validas = [c for c in raw_header if pd.notna(c) and str(c).strip()]

    if len(cells_validas) < 2:
        headers = _split_merged_header(raw_header, ncols)
    else:
        # NÃO dividimos cells merged aqui — isso quebraria o alinhamento com
        # as colunas de dados (que não foram merged). O pós-processamento
        # abaixo extrai data+histórico de uma célula colada.
        headers = [str(c).strip() if pd.notna(c) and str(c).strip() else f'col_{j}'
                   for j, c in enumerate(raw_header)]
        if len(headers) < ncols:
            headers += [f'col_{i}' for i in range(len(headers), ncols)]
        headers = headers[:ncols]

    df = data_rows.copy()
    df.columns = headers
    df = df.reset_index(drop=True).fillna('')

    # Pós-processamento BRB legacy: cabeçalho merged tipo "Data Histórico" e
    # cells de dados com "DD/MM/AA HISTÓRICO" coladas. Splitamos a coluna
    # criando 'Histórico' nova logo após 'Data'.
    col_data = None
    for c in df.columns:
        nome = str(c).strip().lower()
        if nome == 'data' or len(_HEADER_KEYS.findall(nome)) >= 2:
            col_data = c
            break
    col_hist = next((c for c in df.columns
                     if re.search(r'hist[óo]rico|descri[çc][ãa]o', str(c), re.IGNORECASE)
                     and c != col_data), None)
    if col_data is not None and col_hist is None and len(_HEADER_KEYS.findall(str(col_data).lower())) >= 2:
        # Renomeia 'Data Histórico' para 'Data' e cria coluna 'Histórico' adjacente
        idx = list(df.columns).index(col_data)
        cols = list(df.columns)
        cols[idx] = 'Data'
        df.columns = cols
        col_data = 'Data'
        col_hist = 'Histórico'
        df.insert(idx + 1, col_hist, '')
    if col_data is not None:
        rx_data = re.compile(r'^(\d{1,2}[/\-.]\d{1,2}[/\-.]\d{2,4})\s+(.*)$', re.DOTALL)
        for i in df.index:
            cell = str(df.at[i, col_data]).strip()
            m = rx_data.match(cell)
            if m:
                df.at[i, col_data] = m.group(1)
                resto = m.group(2).strip()
                if col_hist is not None and resto:
                    atual = str(df.at[i, col_hist]).strip()
                    df.at[i, col_hist] = resto if not atual else f"{resto} {atual}".strip()

        # Trunca ao primeiro Data inválido após início (corta rodapés tipo
        # "Mensagem Institucional" no AQUAFAN ou "SALDO POUPANCA SALARIO" no BRB_610).
        rx_so_data = re.compile(r'^\d{1,2}[/\-.]\d{1,2}[/\-.]\d{2,4}$|^\d{4}-\d{2}-\d{2}')
        encontrou_dado = False
        cut_idx = None
        for i in df.index:
            cell = str(df.at[i, col_data]).strip()
            valido = bool(rx_so_data.match(cell))
            if valido:
                encontrou_dado = True
            elif encontrou_dado:
                cut_idx = i
                break
        if cut_idx is not None:
            df = df.iloc[:cut_idx].reset_index(drop=True)

    # Resolve ambiguidade do nome 'Lançamento': em BRB legacy é coluna monetária,
    # em BB/outros é descrição. Decide pelo conteúdo: se >70% das cells parecem
    # valor monetário, renomeia para 'Valor'.
    rx_money = re.compile(r'^[+-]?\s*[\d.,]+[+-]?$')
    for c in list(df.columns):
        if str(c).strip().lower() in ('lançamento', 'lancamento'):
            sample = [str(v).strip() for v in df[c].tolist() if str(v).strip()]
            if not sample:
                continue
            hits = sum(1 for v in sample if rx_money.match(v))
            if hits / len(sample) >= 0.7:
                cols = list(df.columns)
                cols[cols.index(c)] = 'Valor'
                df.columns = cols

    return df


def ler_df(filepath: str, filename: str) -> pd.DataFrame:
    """
    Lê CSV, XLSX, OFX ou PDF para DataFrame com todas as colunas como string.
    Aplica fillna('') para eliminar float NaN antes do pipeline.
    CSV: testa encodings utf-8, latin-1, cp1252 em sequência.
    Retorna DataFrame vazio em caso de falha.
    """
    ext = filename.rsplit('.', 1)[-1].lower()
    try:
        if ext == 'csv':
            for enc in ['utf-8', 'latin-1', 'cp1252']:
                try:
                    df = pd.read_csv(filepath, encoding=enc, on_bad_lines='skip', dtype=str)
                    return df.fillna('')
                except:
                    pass
        elif ext in ['xls', 'xlsx']:
            df = _read_excel_smart_header(filepath)
            return df.fillna('')
        elif ext == 'ofx':
            with open(filepath, 'rb') as f:
                ofx = OfxParser.parse(f)
            rows = []
            for txn in ofx.account.statement.transactions:
                rows.append({
                    'Data': txn.date.strftime('%d/%m/%Y'),
                    'Valor': str(txn.amount),
                    'Descricao': txn.memo or txn.payee or '',
                    'Tipo': 'entrada' if txn.amount >= 0 else 'saida',
                })
            if not rows:
                return pd.DataFrame()
            return pd.DataFrame(rows).astype(str)
        elif ext == 'pdf':
            if not HAS_PDF_EXTRACTOR:
                print(f"[LEITURA] PDF {filename}: extrator indisponível")
                return pd.DataFrame()
            df, warnings = _extrair_pdf(filepath, filename)
            # Persiste warnings na sessão para inspeção via API
            SESSION.setdefault('pdf_warnings', {})[filename] = warnings
            for w in warnings:
                print(f"[PDF] {w}")
            if df is None or df.empty:
                return pd.DataFrame()
            return df.fillna('').astype(str)
    except Exception as e:
        print(f"[LEITURA] Falha em {filename}: {e}")
        if ext == 'pdf':
            SESSION.setdefault('pdf_warnings', {})[filename] = [
                f'Falha crítica na extração: {e}'
            ]
    return pd.DataFrame()

def _md5(filepath: Path) -> str:
    """MD5 de arquivo em chunks — usado para deduplicação por conteúdo."""
    h = hashlib.md5()
    with open(filepath, 'rb') as fh:
        for chunk in iter(lambda: fh.read(65536), b''):
            h.update(chunk)
    return h.hexdigest()


# ============================================================
# HELPERS — DETECÇÃO DE UNIDADE
# ============================================================

def encontrar_unidade(text: str) -> tuple[str | None, int, str]:
    """
    Identifica unidade de negócio pelo nome do arquivo.

    Três camadas:
      1. Padrões aprendidos (fn_patterns em config.json)
      2. Keywords hardcoded do Grupo Manta/GN
      3. Fuzzy matching contra cadastro de unidades (limiar 70%)

    Returns: (unit_id | None, confianca 0-100, metodo)
    """
    tu = text.upper()

    for pattern, uid in CFG.get('fn_patterns', {}).items():
        if pattern.upper() in tu:
            return uid, 95, 'aprendido'

    PATTERNS = [
        (('303SW', '303'),                          ('MN303', 90)),
        (('610',),                                  ('MN610', 90)),
        (('702',),                                  ('MN702', 90)),
        (('712',),                                  ('MN712', 90)),
        (('LUDIKA', 'LUDICA'),                      ('LD1',   100)),
        (('GN', 'ADMINISTRATIVO', 'CSC'),           ('GN',    88)),
        (('MANTA-VP', 'MANTAVP', '_VP_', '-VP-'),   ('MNVP',  92)),
        (('MANTA-JB', 'MANTAJB', '_JB_', '-JB-'),   ('MNJB',  92)),
        (('RAIACLUBE', 'RAIA-CLUBE', 'RAIACLUB'),   ('RC608', 88)),
        (('ASTCU',),                                ('RCA',   92)),
    ]
    for keywords, (uid, conf) in PATTERNS:
        for kw in keywords:
            if kw in tu:
                return uid, conf, 'exata'

    best, bsc = None, 0
    for u in CFG['unidades']:
        for field in [u['desc_unidade'], u['marca'], u['id']]:
            sc = difflib.SequenceMatcher(None, tu, field.upper()).ratio() * 100
            if sc > bsc:
                bsc = sc; best = u['id']
    if bsc >= 70:
        return best, int(bsc), 'fuzzy'

    return None, 0, 'nao_encontrado'

def construir_preview(arq: dict) -> dict:
    """
    Sumário rápido de um arquivo sem processamento completo.
    Resultado cacheado em SESSION['previews'].
    """
    fn = arq['filename']
    df = ler_df(arq['path'], fn)
    if df.empty:
        return {'ok': False, 'rows': 0, 'msg': 'Arquivo vazio ou ilegível',
                'sum_entrada': 0, 'sum_saida': 0, 'periodo': '-', 'colunas': []}

    sm  = detectar_mapa(list(df.columns))
    inv = inv_mapa(sm)
    entradas, saidas, datas = [], [], []

    for _, row in df.iterrows():
        rd   = row.to_dict()
        tipo, _ = detectar_tipo(rd, inv)
        if inv.get('Debito') and inv.get('Credito'):
            d = parse_valor(rd.get(inv['Debito']))
            c = parse_valor(rd.get(inv['Credito']))
            if tipo == 'saida'   and d: saidas.append(abs(d))
            if tipo == 'entrada' and c: entradas.append(abs(c))
        elif inv.get('Valor'):
            n = parse_valor(rd.get(inv['Valor']))
            if n is not None:
                (saidas if tipo == 'saida' else entradas).append(abs(n))
        if inv.get('Data'):
            dt = parse_data(rd.get(inv['Data']))
            if dt and len(dt) == 10: datas.append(dt)

    return {
        'ok': True, 'rows': len(df),
        'sum_entrada':  round(sum(entradas), 2),
        'sum_saida':    round(sum(saidas),   2),
        'periodo':      f"{min(datas)} → {max(datas)}" if datas else '-',
        'colunas':      list(df.columns),
        'mapeadas':     {col: anc for col, anc in sm.items() if anc},
        'nao_mapeadas': [col for col, anc in sm.items() if not anc],
    }


# ============================================================
# ROTAS — RENDERIZAÇÃO
# ============================================================

@app.route('/')
def index():
    return render_template('index.html')


# ============================================================
# ROTAS — UPLOAD
# ============================================================

@app.route('/api/upload', methods=['POST'])
def upload():
    """
    Recebe arquivos com três camadas de validação:
      1. Extensão: apenas CSV, XLS, XLSX, OFX, PDF
      2. Nome duplicado na sessão
      3. Conteúdo duplicado (MD5)
    Retorna listas 'uploaded' e 'rejected' com motivos.
    """
    files   = request.files.getlist('files')
    uploaded, rejected = [], []
    nomes_existentes  = {a['filename'] for a in SESSION['arquivos']}
    hashes_existentes = {a['hash']: a['filename']
                         for a in SESSION['arquivos'] if 'hash' in a}

    for f in files:
        if not f.filename: continue
        ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
        if ext == 'pdf' and not HAS_PDF_EXTRACTOR:
            rejected.append({'filename': f.filename,
                             'motivo': 'PDF: instale pdfplumber/pymupdf para habilitar'})
            continue
        if ext not in ALLOWED_EXT:
            rejected.append({'filename': f.filename,
                             'motivo': f'Extensão .{ext} não suportada'})
            continue
        fn = secure_filename(f.filename)
        if fn in nomes_existentes:
            rejected.append({'filename': f.filename,
                             'motivo': 'Arquivo já carregado nesta sessão (mesmo nome)'})
            continue
        fp = UPLOAD_FOLDER / fn
        f.save(fp)
        file_hash = _md5(fp)
        if file_hash in hashes_existentes:
            fp.unlink(missing_ok=True)
            rejected.append({'filename': f.filename,
                             'motivo': f'Conteúdo idêntico ao arquivo "{hashes_existentes[file_hash]}"'})
            continue
        entry = {'filename': fn, 'size': fp.stat().st_size,
                 'path': str(fp), 'hash': file_hash}
        SESSION['arquivos'].append(entry)
        uploaded.append(entry)
        nomes_existentes.add(fn)
        hashes_existentes[file_hash] = fn

    return jsonify({'ok': True, 'uploaded': uploaded, 'rejected': rejected,
                    'total': len(SESSION['arquivos'])})

@app.route('/api/arquivos')
def listar_arquivos():
    return jsonify({'arquivos': SESSION['arquivos']})

@app.route('/api/pdf-warnings')
def pdf_warnings():
    """Retorna warnings/quarentena gerados pela extração de PDFs."""
    return jsonify({
        'available': HAS_PDF_EXTRACTOR,
        'warnings': SESSION.get('pdf_warnings', {})
    })

@app.route('/api/remover-arquivo/<filename>', methods=['DELETE'])
def remover_arquivo(filename):
    """Remove arquivo da sessão e do disco. Limpa estados associados."""
    fp = UPLOAD_FOLDER / filename
    try: fp.unlink(missing_ok=True)
    except: pass
    SESSION['arquivos'] = [a for a in SESSION['arquivos'] if a['filename'] != filename]
    for chave in ('previews', 'doc_verificados', 'deteccao_cab', 'cruzamento',
                  'conta_por_arquivo', 'conta_por_arquivo_meta'):
        SESSION[chave].pop(filename, None)
    return jsonify({'ok': True})


# ============================================================
# ROTAS — PREVIEW
# ============================================================

@app.route('/api/preview/<filename>')
def preview_arquivo(filename):
    """Sumário de arquivo com cache em SESSION['previews']."""
    arq = next((a for a in SESSION['arquivos'] if a['filename'] == filename), None)
    if not arq: return jsonify({'ok': False, 'msg': 'Arquivo não encontrado'}), 404
    if filename not in SESSION['previews']:
        SESSION['previews'][filename] = construir_preview(arq)
    return jsonify(SESSION['previews'][filename])


# ============================================================
# ROTAS — VERIFICAÇÃO DE UNIDADES
# ============================================================

@app.route('/api/verificar-documentos')
def verificar_documentos():
    """Lista arquivos com dupla identificação (cabeçalho × nome do arquivo).

    Para cada arquivo, roda o detector de cabeçalho (se ainda não rodou neste
    ciclo), o encontrar_unidade (controle secundário) e validar_cruzado.
    A resposta é estruturada em dois blocos (cabecalho, nome_arquivo) mais
    o banner de status do cruzamento.
    """
    documentos = []
    contas_cfg = CFG.get('contas_bancarias', [])

    for arq in SESSION['arquivos']:
        fn   = arq['filename']
        path = arq['path']

        deteccao = SESSION['deteccao_cab'].get(fn)
        if deteccao is None:
            deteccao = detectar_cabecalho(path, fn)
            SESSION['deteccao_cab'][fn] = deteccao

        conta_id_cab, conta_conf, conta_motivo = (
            resolver_conta(deteccao, contas_cfg) if deteccao else (None, 0, 'sem_cabecalho')
        )
        conta_obj = next((c for c in contas_cfg if c['id'] == conta_id_cab), None) if conta_id_cab else None

        uid_nome, conf_nome, metodo_nome = encontrar_unidade(fn)
        unit_id_nome = uid_nome if conf_nome >= 80 else None

        cruzamento = validar_cruzado(conta_id_cab, unit_id_nome, contas_cfg)
        SESSION['cruzamento'][fn] = cruzamento

        confirmacao = SESSION['conta_por_arquivo'].get(fn)
        meta_conf   = SESSION['conta_por_arquivo_meta'].get(fn)

        documentos.append({
            'filename': fn,
            'preview':  SESSION['previews'].get(fn, {}),
            'cabecalho': {
                'banco_detectado': deteccao.get('banco_detectado') if deteccao else None,
                'banco_codigo':    deteccao.get('banco_codigo')    if deteccao else None,
                'agencia':         deteccao.get('agencia')         if deteccao else None,
                'conta':           deteccao.get('conta')           if deteccao else None,
                'cnpj_titular':    deteccao.get('cnpj_titular')    if deteccao else None,
                'conta_id_resolvida':  conta_id_cab,
                'conta_desc_resolvida': (
                    f"{conta_obj['banco_nome']} — ag {conta_obj.get('agencia') or '-'} / "
                    f"conta {conta_obj.get('conta') or '-'}"
                ) if conta_obj else None,
                'unit_id_resolvida':   conta_obj.get('unit_id') if conta_obj else None,
                'conta_confianca':     conta_conf,
                'conta_motivo':        conta_motivo,
            },
            'nome_arquivo': {
                'unit_id_sugerido': uid_nome,
                'confianca':        conf_nome,
                'metodo':           metodo_nome,
            },
            'cruzamento':      cruzamento,
            'confirmacao':     confirmacao,
            'confirmacao_meta': meta_conf,
            # Legado v8 — conservado para não quebrar UI atual enquanto migra.
            'detectada':  uid_nome,
            'confianca':  conf_nome,
            'metodo':     metodo_nome,
            'confirmada': SESSION['doc_verificados'].get(fn),
            'status': (
                'verificado' if cruzamento['status'] in ('concordam', 'apenas_cabecalho')
                              or confirmacao or cruzamento['status'] == 'modo_legado'
                else 'pendente'
            ),
        })
    return jsonify({'documentos': documentos,
                    'modo_legado': not CFG.get('contas_bancarias')})

@app.route('/api/confirmar-unidade', methods=['POST'])
def confirmar_unidade():
    """
    Registra confirmação manual de unidade.
    Aprendizado: salva stem do filename em fn_patterns para execuções futuras.
    """
    data = request.get_json()
    fn, uid = data.get('filename'), data.get('unit_id')
    if not uid:
        SESSION['doc_verificados'].pop(fn, None)
        return jsonify({'ok': True})
    if not any(u['id'] == uid for u in CFG['unidades']):
        return jsonify({'ok': False, 'msg': 'Unidade não encontrada'}), 404
    SESSION['doc_verificados'][fn] = uid
    stem = Path(fn).stem[:25].upper()
    CFG['fn_patterns'][stem] = uid
    save_config(CFG)
    return jsonify({'ok': True})


# ============================================================
# HELPERS v9 — DECISÃO DE ATRIBUIÇÃO CONTA / UNIDADE
# ============================================================

def _unit_id_from_conta(conta_id: str | None) -> str | None:
    if not conta_id:
        return None
    for c in CFG.get('contas_bancarias', []):
        if c.get('id') == conta_id:
            return c.get('unit_id')
    return None


def _resolver_atribuicao(fn: str, cruzamento: dict, uid_nome: str | None,
                          cfg: dict) -> tuple[str | None, str | None, str, str]:
    """Decide a atribuição final (unit_id, conta_id, rastreab, metodo) para um arquivo.

    Consulta SESSION['conta_por_arquivo'] para casos que exigem decisão humana.
    Em casos de bloqueio (conflito/apenas_nome/nenhum sem confirmação), registra
    em SESSION['arquivos_bloqueados'] e devolve (None, None, ...) — o loop de
    processamento deve pular arquivos nessa situação.

    Retorna: (unit_id, conta_id, confiab_rastreab, metodo_atribuicao)
    """
    status           = cruzamento['status']
    confirmacao      = SESSION['conta_por_arquivo'].get(fn)
    meta_confirmacao = SESSION['conta_por_arquivo_meta'].get(fn, {})

    # Casos limpos — cabeçalho decide (com ou sem concordância do nome)
    if status in ('concordam', 'apenas_cabecalho'):
        return (cruzamento['unit_id'], cruzamento['conta_id'],
                'ALTA', cruzamento['metodo'])

    # Modo legado — preserva comportamento v8
    if status == 'modo_legado':
        uid = SESSION['doc_verificados'].get(fn) or uid_nome
        return uid, None, 'BAIXA', 'modo_legado'

    # Conflito, apenas_nome, nenhum: requer confirmação humana.
    if confirmacao:
        unit_id_final = _unit_id_from_conta(confirmacao)
        rastreab      = 'MEDIA'
        metodo        = f"{status}_confirmado_manual:{confirmacao}"

        SESSION['audit_log'].append({
            'ts':        datetime.now().isoformat(timespec='seconds'),
            'filename':  fn,
            'acao':      f'{status}_resolvido_manualmente',
            'detalhes': {
                'metodo_original':      cruzamento['metodo'],
                'conta_cabecalho':      cruzamento.get('conta_id'),
                'unit_id_cabecalho':    cruzamento.get('unit_id_cabecalho'),
                'unit_id_nome_arquivo': cruzamento.get('unit_id_nome_arquivo'),
                'escolha_humana':       confirmacao,
                'motivo':               meta_confirmacao.get('motivo', ''),
            },
        })
        return unit_id_final, confirmacao, rastreab, metodo

    # Sem confirmação: bloqueia.
    motivo_msg = {
        'conflito':    f"Conflito entre cabeçalho ({cruzamento.get('unit_id_cabecalho')}) "
                       f"e nome do arquivo ({cruzamento.get('unit_id_nome_arquivo')}). "
                       f"Confirme manualmente antes de processar.",
        'apenas_nome': f"Nome do arquivo sugere {cruzamento.get('unit_id_nome_arquivo')} "
                       f"mas o cabeçalho não foi reconhecido. Confirme a conta manualmente.",
        'nenhum':      "Nem cabeçalho nem nome do arquivo identificam a conta. "
                       "Cadastre a conta em Configurações e confirme o arquivo.",
    }.get(status, f'Status não processável: {status}')

    SESSION['arquivos_bloqueados'].append({
        'filename':              fn,
        'status':                status,
        'conta_id_cabecalho':    cruzamento.get('conta_id'),
        'unit_id_cabecalho':     cruzamento.get('unit_id_cabecalho'),
        'unit_id_nome_arquivo': cruzamento.get('unit_id_nome_arquivo'),
        'msg':                   motivo_msg,
    })
    return None, None, 'BAIXA', cruzamento['metodo']


# ============================================================
# ROTAS — PROCESSAMENTO PRINCIPAL
# ============================================================

@app.route('/api/processar', methods=['POST'])
def processar():
    """
    Pipeline completo de extração e classificação.

    Passo 1 — Schema global: detecta mapeamento coluna→anchor em todos os arquivos.
    Passo 2 — Linha a linha: extrai campos, aplica fallbacks B1/B2/B3, classifica.

    Progresso disponível via GET /api/progresso (polling).
    """
    if not SESSION['arquivos']:
        return jsonify({'ok': False, 'msg': 'Nenhum arquivo carregado'}), 400

    SESSION['progresso'] = {'pct': 2, 'msg': 'Detectando schema...', 'ativo': True}

    # Passo 1: schema global (usa o arquivo com mais anchors obrigatórios como referência)
    all_file_schemas = {}
    richest_cols, richest_score = [], 0
    for arq in SESSION['arquivos']:
        df = ler_df(arq['path'], arq['filename'])
        if df.empty: continue
        cols  = list(df.columns)
        sm    = detectar_mapa(cols)
        score = sum(1 for v in sm.values() if v in ANCHORS_OBRIGATORIOS)
        all_file_schemas[arq['filename']] = sm
        if score > richest_score:
            richest_score = score; richest_cols = cols

    global_schema = detectar_mapa(richest_cols) if richest_cols else {}

    # Persiste novas descobertas no de-para
    updated = False
    for col, anc in global_schema.items():
        if anc and col not in CFG['depara']:
            CFG['depara'][col] = anc; updated = True
    if updated: save_config(CFG)
    SESSION['schema_map'] = global_schema

    # Passo 2: processamento linha a linha
    todos = []
    total = len(SESSION['arquivos'])

    # v9: estado de detecção e bloqueios por arquivo é reiniciado a cada processar().
    SESSION['deteccao_cab'] = {}
    SESSION['cruzamento']   = {}
    SESSION['arquivos_bloqueados'] = []

    erros_arquivos = []
    for idx, arq in enumerate(SESSION['arquivos']):
        fn  = arq['filename']
        pct = 5 + int((idx / total) * 90)
        SESSION['progresso'] = {'pct': pct, 'msg': f'Processando {fn} ({idx+1}/{total})...', 'ativo': True}

        try:
            # ── v9: Identificação por cabeçalho (fonte primária) ──────────
            deteccao_cab = detectar_cabecalho(arq['path'], fn)
            SESSION['deteccao_cab'][fn] = deteccao_cab

            conta_id_cab = None
            if deteccao_cab:
                conta_id_cab, _, _ = resolver_conta(deteccao_cab, CFG['contas_bancarias'])

            # ── v9: Controle secundário pelo nome do arquivo (NUNCA decide sozinho) ──
            uid_nome, conf_nome, _ = encontrar_unidade(fn)
            unit_id_nome = uid_nome if conf_nome >= 80 else None

            # ── v9: Cruzamento e matriz de decisão ─────────────────────────
            cruzamento = validar_cruzado(conta_id_cab, unit_id_nome, CFG['contas_bancarias'])
            SESSION['cruzamento'][fn] = cruzamento

            uid, conta_id_final, rastreab, metodo_atrib = _resolver_atribuicao(
                fn, cruzamento, uid_nome, CFG
            )

            # Arquivo bloqueado → não entra em lançamentos.
            if uid is None and conta_id_final is None and cruzamento['status'] != 'modo_legado':
                continue

            unidade_info = next((u for u in CFG['unidades'] if u['id'] == uid), None) if uid else None
            marca   = unidade_info['marca']        if unidade_info else 'N/D'
            unidade = unidade_info['desc_unidade'] if unidade_info else 'N/D'

            # Em modo_legado, preserva o comportamento v8 de consultar doc_verificados
            # apenas para o cálculo de confiabilidade da unidade (classificar).
            unit_conf_para_classificar      = 100 if rastreab != 'BAIXA' else conf_nome
            unit_confirmada_para_classificar = (rastreab != 'BAIXA' or
                                                 bool(SESSION['doc_verificados'].get(fn)))

            df = ler_df(arq['path'], fn)
            if df.empty: continue

            file_schema = all_file_schemas.get(fn) or detectar_mapa(list(df.columns))
            inv         = inv_mapa(file_schema)

            for i, (_, row) in enumerate(df.iterrows()):
                rd     = row.to_dict()
                extras = {f'extra_{col}': val for col, val in rd.items()
                          if not file_schema.get(col)}

                anchors_p = sum(1 for a in ['Data','Valor','Descricao']
                                if rd.get(inv.get(a,''),'') not in ('', None))
                extras_u  = sum(1 for k, v in extras.items()
                                if v not in ('', None) and 'Unnamed' not in k
                                and k not in ('extra_Nosso Número',))
                if anchors_p == 0 and extras_u == 0:
                    continue

                desc_raw = str(rd.get(inv.get('Descricao',''),'') if inv.get('Descricao') else '').strip()
                data_raw = str(rd.get(inv.get('Data',''),'') if inv.get('Data') else '').strip()
                _LIXO_PDF = {'saldo anterior', 'saldo do dia', 'data', 'descrição', 'descricao',
                             'valor', 'saldo', 'lançamentos', 'lancamentos', 'doc', 'mensagem institucional'}
                if desc_raw.lower() in _LIXO_PDF or (not data_raw and not desc_raw):
                    continue

                tipo, tipo_conf = detectar_tipo(rd, inv)

                if inv.get('Debito') and inv.get('Credito'):
                    d = parse_valor(rd.get(inv['Debito']))
                    c = parse_valor(rd.get(inv['Credito']))
                    valor_abs = (abs(d) if tipo == 'saida' and d is not None
                                 else abs(c) if c is not None
                                 else abs(d) if d is not None else None)
                elif inv.get('Valor'):
                    n = parse_valor(rd.get(inv['Valor']))
                    valor_abs = abs(n) if n is not None else None
                else:
                    valor_abs = None

                if valor_abs is None:
                    sa_raw = extras.get('extra_Saldo antes') or extras.get('extra_Saldo Antes')
                    sd_raw = extras.get('extra_Saldo depois') or extras.get('extra_Saldo Depois')
                    if sa_raw and sd_raw:
                        sa = parse_valor(str(sa_raw).replace('R$','').replace('.','').replace(',','.'))
                        sd = parse_valor(str(sd_raw).replace('R$','').replace('.','').replace(',','.'))
                        if sa is not None and sd is not None:
                            v = round(abs(sd - sa), 2)
                            valor_abs = v if v > 0 else None
                    if tipo == 'indefinido':
                        mov = str(extras.get('extra_Movimentação') or
                                   extras.get('extra_Movimentacao', '')).strip().upper()
                        if mov in ('CRÉDITO','CREDITO','CREDIT','CR'):
                            tipo = 'entrada'; tipo_conf = 'alta'
                        elif mov in ('DÉBITO','DEBITO','DEBIT','DB'):
                            tipo = 'saida'; tipo_conf = 'alta'

                descricao_raw = str(rd.get(inv['Descricao'],'') if inv.get('Descricao') else '').strip()

                if not descricao_raw:
                    t     = str(extras.get('extra_Transação') or extras.get('extra_Transacao') or '').strip()
                    ident = str(extras.get('extra_Identificação') or extras.get('extra_Identificacao') or '').strip()
                    partes = [p for p in [t, ident] if p and p.lower() not in ('nan','none','')]
                    if partes: descricao_raw = ' | '.join(partes)

                if not descricao_raw:
                    cat = str(extras.get('extra_Tipo') or '').strip()
                    dst = str(extras.get('extra_Destino') or extras.get('extra_Origem') or '').strip()
                    partes = [p for p in [cat, dst]
                              if p and p.lower() not in ('nan','none','desconhecido','')]
                    if partes: descricao_raw = ' | '.join(partes)

                if inv.get('Data'):
                    data_iso, data_raw = parse_data(rd.get(inv['Data']), preservar_raw=True)
                else:
                    data_iso, data_raw = None, ''

                row_norm = {
                    'Data':         data_iso,
                    'Data_Raw':     data_raw,
                    'Valor':        valor_abs,
                    'Tipo':         tipo,
                    'Descricao':    descricao_raw,
                    'Conta':        str(rd.get(inv['Conta'],        '') if inv.get('Conta')        else '').strip(),
                    'Banco':        str(rd.get(inv['Banco'],        '') if inv.get('Banco')        else '').strip(),
                    'CNPJ':         str(rd.get(inv['CNPJ'],         '') if inv.get('CNPJ')         else '').strip(),
                    'Centro_Custo': str(rd.get(inv['Centro_Custo'], '') if inv.get('Centro_Custo') else '').strip(),
                }

                grupo, conf, issues = classificar(row_norm, unit_conf_para_classificar,
                                                   unit_confirmada_para_classificar)
                # v9: anexa issue de rastreabilidade ao lançamento quando aplicável.
                if cruzamento.get('issue') and cruzamento['issue'] not in issues:
                    issues.append(cruzamento['issue'])
                    if grupo == 'A':
                        grupo = 'B'
                        conf  = 'BAIXA' if conf == 'ALTA' else conf

                todos.append({
                    'id': f"{fn}::{i}", 'arquivo': fn,
                    'unidade_id': uid or '', 'marca': marca, 'unidade': unidade,
                    'conta_id': conta_id_final,
                    'metodo_atribuicao': metodo_atrib,
                    'Confiab_Rastreabilidade': rastreab,
                    'grupo': grupo, 'confiabilidade': conf, 'issues': issues,
                    'tipo_conf': tipo_conf, 'status': 'pendente',
                    **row_norm, 'extras': extras,
                })
        except Exception as e:
            erros_arquivos.append(fn)
            print(f"[PROCESSAR] Erro em {fn}: {e}")

    SESSION['lancamentos'] = todos
    SESSION['processado']  = True
    SESSION['progresso']   = {'pct': 100, 'msg': 'Concluído', 'ativo': False}

    ga = sum(1 for l in todos if l['grupo'] == 'A')
    gb = sum(1 for l in todos if l['grupo'] == 'B')
    resp = {'ok': True, 'total': len(todos), 'grupo_a': ga, 'grupo_b': gb,
            'schema_map': {k: v for k, v in global_schema.items() if v}}
    if erros_arquivos:
        resp['erros'] = erros_arquivos
        resp['msg'] = f'{len(erros_arquivos)} arquivo(s) com erro: {", ".join(erros_arquivos)}'
    if SESSION['arquivos_bloqueados']:
        resp['bloqueados'] = SESSION['arquivos_bloqueados']
        nomes = ', '.join(b['filename'] for b in SESSION['arquivos_bloqueados'])
        resp['msg'] = (
            f"{len(SESSION['arquivos_bloqueados'])} arquivo(s) bloqueado(s) por "
            f"conflito/ausência de identificação: {nomes}. "
            f"Resolva em 'Verificação' antes de reprocessar."
        )
    return jsonify(resp)

@app.route('/api/progresso')
def progresso():
    return jsonify(SESSION['progresso'])


# ============================================================
# ROTAS — LANÇAMENTOS (paginado)
# ============================================================

@app.route('/api/lancamentos')
def listar_lancamentos():
    """
    Lista lançamentos paginados com filtros opcionais.
    Query params: grupo, status, page (default 1), per_page (default 50).
    """
    grupo    = request.args.get('grupo')
    status_f = request.args.get('status')
    page     = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 50))

    items = SESSION['lancamentos']
    if grupo:    items = [l for l in items if l['grupo']  == grupo]
    if status_f: items = [l for l in items if l['status'] == status_f]

    total = len(items)
    start = (page - 1) * per_page
    return jsonify({
        'items':         items[start:start + per_page],
        'total':         total,
        'page':          page,
        'pages':         max(1, (total + per_page - 1) // per_page),
        'grupo_a_total': sum(1 for l in SESSION['lancamentos'] if l['grupo'] == 'A'),
        'grupo_b_total': sum(1 for l in SESSION['lancamentos'] if l['grupo'] == 'B'),
    })


# ============================================================
# ROTAS — AÇÕES SOBRE LANÇAMENTOS
# ============================================================

@app.route('/api/acao', methods=['POST'])
def acao():
    """
    Ações individuais ou em lote sobre lançamentos.

    acao='confirmar_grupo_a' → confirma todos os pendentes do Grupo A
    acao='confirmar'         → confirma lançamento pelo id
    acao='excluir'           → exclui lançamento pelo id
    acao='editar'            → atualiza campos e reclassifica; promove para A se possível
    """
    data      = request.get_json()
    lid       = data.get('id')
    acao_tipo = data.get('acao')
    campos    = data.get('campos', {})

    if acao_tipo == 'confirmar_grupo_a':
        n = 0
        for l in SESSION['lancamentos']:
            if l['grupo'] == 'A' and l['status'] == 'pendente':
                l['status'] = 'confirmado'; n += 1
        return jsonify({'ok': True, 'n': n})

    for l in SESSION['lancamentos']:
        if l['id'] == lid:
            if acao_tipo == 'confirmar':
                l['status'] = 'confirmado'
            elif acao_tipo == 'excluir':
                l['status'] = 'excluido'
            elif acao_tipo == 'editar':
                for k, v in campos.items():
                    if k in l: l[k] = v
                uid_conf = bool(SESSION['doc_verificados'].get(l['arquivo']))
                uconf    = 100 if uid_conf else encontrar_unidade(l['arquivo'])[1]
                g, c, iss = classificar(l, uconf, uid_conf)
                l['grupo'] = g; l['confiabilidade'] = c; l['issues'] = iss
                if g == 'A': l['status'] = 'confirmado'
            break

    return jsonify({'ok': True})


@app.route('/api/acao-bloco', methods=['POST'])
def acao_bloco():
    """
    Ação em lote sobre subconjuntos do Grupo B.

    Permite resolver rapidamente grupos homogêneos sem clicar item a item.
    Três eixos de agrupamento:

      eixo='issue'    → todos com o mesmo tipo de problema
                        Ex: valor='descricao_ausente' confirma os 3.520 do Cora de uma vez
      eixo='arquivo'  → todos do mesmo arquivo de extrato
      eixo='unidade'  → todos da mesma unidade de negócio

    Body: {acao: 'confirmar'|'excluir', eixo: 'issue'|'arquivo'|'unidade', valor: str}
    Returns: n (afetados) + resumo atualizado de contagens
    """
    data  = request.get_json()
    acao  = data.get('acao')
    eixo  = data.get('eixo')
    valor = data.get('valor', '')

    if acao not in ('confirmar', 'excluir'):
        return jsonify({'ok': False, 'msg': 'Ação inválida. Use confirmar ou excluir'}), 400
    if eixo not in ('issue', 'arquivo', 'unidade'):
        return jsonify({'ok': False, 'msg': 'Eixo inválido. Use issue, arquivo ou unidade'}), 400

    n = 0
    for l in SESSION['lancamentos']:
        if l['grupo'] != 'B' or l['status'] != 'pendente':
            continue
        match = (
            (eixo == 'issue'   and valor in l.get('issues', [])) or
            (eixo == 'arquivo' and l['arquivo'] == valor)         or
            (eixo == 'unidade' and l['unidade'] == valor)
        )
        if match:
            l['status'] = acao + 'do'  # 'confirmado' ou 'excluido'
            n += 1

    ls = SESSION['lancamentos']
    return jsonify({
        'ok': True, 'n': n,
        'resumo': {
            'pendentes':   sum(1 for l in ls if l['status'] == 'pendente'),
            'confirmados': sum(1 for l in ls if l['status'] == 'confirmado'),
            'excluidos':   sum(1 for l in ls if l['status'] == 'excluido'),
            'b_pendentes': sum(1 for l in ls if l['grupo'] == 'B' and l['status'] == 'pendente'),
        }
    })


@app.route('/api/bloco-resumo')
def bloco_resumo():
    """
    Subgrupos do Grupo B para o painel de ações em lote.
    Retorna contagens por issue, arquivo e unidade — alimenta a UI da Tab 4.
    """
    pendentes_b = [l for l in SESSION['lancamentos']
                   if l['grupo'] == 'B' and l['status'] == 'pendente']

    ISSUE_LABELS = {
        'data_ausente': 'Data ausente', 'valor_ausente': 'Valor ausente',
        'descricao_ausente': 'Descrição ausente', 'tipo_indefinido': 'Tipo indefinido',
        'unidade_incerta': 'Unidade incerta',
    }

    issue_counts = {}
    for l in pendentes_b:
        for iss in l.get('issues', []):
            issue_counts[iss] = issue_counts.get(iss, 0) + 1

    arq_counts = {}
    for l in pendentes_b:
        arq_counts[l['arquivo']] = arq_counts.get(l['arquivo'], 0) + 1

    uni_counts = {}
    for l in pendentes_b:
        uni_counts[l['unidade']] = uni_counts.get(l['unidade'], 0) + 1

    return jsonify({
        'total_pendentes': len(pendentes_b),
        'por_issue':   [{'valor': iss, 'label': ISSUE_LABELS.get(iss, iss), 'count': cnt}
                        for iss, cnt in sorted(issue_counts.items(), key=lambda x: -x[1])],
        'por_arquivo': [{'valor': arq, 'count': cnt}
                        for arq, cnt in sorted(arq_counts.items(),  key=lambda x: -x[1])],
        'por_unidade': [{'valor': uni, 'count': cnt}
                        for uni, cnt in sorted(uni_counts.items(),  key=lambda x: -x[1])],
    })


# ============================================================
# ROTAS — RESUMO E GATES
# ============================================================

@app.route('/api/resumo')
def resumo():
    """Contadores e totais financeiros para a Tab 5."""
    ls      = SESSION['lancamentos']
    nao_excl = [l for l in ls if l['status'] != 'excluido']
    return jsonify({
        'total':        len(ls),
        'confirmados':  sum(1 for l in ls if l['status'] == 'confirmado'),
        'excluidos':    sum(1 for l in ls if l['status'] == 'excluido'),
        'pendentes':    sum(1 for l in ls if l['status'] == 'pendente'),
        'grupo_a':      sum(1 for l in ls if l['grupo']  == 'A'),
        'grupo_b':      sum(1 for l in ls if l['grupo']  == 'B'),
        'b_pendentes':  sum(1 for l in ls if l['grupo']  == 'B' and l['status'] == 'pendente'),
        'sum_entradas': round(sum(l['Valor'] or 0 for l in nao_excl if l['Tipo'] == 'entrada'), 2),
        'sum_saidas':   round(sum(l['Valor'] or 0 for l in nao_excl if l['Tipo'] == 'saida'),   2),
    })

@app.route('/api/gate/<int:step>')
def gate(step):
    """
    Valida pré-condições para avançar entre etapas.
    Gates 1→2, 2→3, 3→4, 4→5 com mensagens de bloqueio ou alerta.
    """
    if step == 1:
        ok = len(SESSION['arquivos']) > 0
        return jsonify({'ok': ok, 'msg': '' if ok else 'Faça upload de pelo menos 1 arquivo'})
    if step == 2:
        modo_legado = not CFG.get('contas_bancarias')
        pendentes = []
        contas_cfg = CFG.get('contas_bancarias', [])
        for a in SESSION['arquivos']:
            fn = a['filename']
            # Garante detecção/cruzamento antes de avaliar.
            deteccao = SESSION['deteccao_cab'].get(fn)
            if deteccao is None:
                deteccao = detectar_cabecalho(a['path'], fn)
                SESSION['deteccao_cab'][fn] = deteccao
            conta_id_cab, _, _ = (resolver_conta(deteccao, contas_cfg)
                                   if deteccao else (None, 0, 'sem_cabecalho'))
            uid_nome, conf_nome, _ = encontrar_unidade(fn)
            unit_id_nome = uid_nome if conf_nome >= 80 else None
            cruz = validar_cruzado(conta_id_cab, unit_id_nome, contas_cfg)
            SESSION['cruzamento'][fn] = cruz

            liberado = (
                cruz['status'] in ('concordam', 'apenas_cabecalho') or
                bool(SESSION['conta_por_arquivo'].get(fn)) or
                cruz['status'] == 'modo_legado' or
                # legacy v8: aceita confirmação de unidade pela UI antiga
                (modo_legado and (SESSION['doc_verificados'].get(fn) or conf_nome >= 80))
            )
            if not liberado:
                pendentes.append({
                    'filename': fn,
                    'status':   cruz['status'],
                    'msg':      cruz.get('issue'),
                })

        msg = ''
        if pendentes:
            msg = (f"{len(pendentes)} arquivo(s) com conta não resolvida — "
                   f"resolva conflitos ou cadastre contas antes de processar.")
        return jsonify({'ok': True, 'pendentes': pendentes, 'msg': msg,
                        'modo_legado': modo_legado})
    if step == 3:
        ok = SESSION['processado'] and len(SESSION['lancamentos']) > 0
        return jsonify({'ok': ok, 'msg': '' if ok else 'Execute o processamento primeiro'})
    if step == 4:
        bp = sum(1 for l in SESSION['lancamentos']
                 if l['grupo'] == 'B' and l['status'] == 'pendente')
        return jsonify({'ok': True, 'b_pendentes': bp,
                        'msg': f'{bp} item(ns) do Grupo B ainda pendentes' if bp else ''})
    return jsonify({'ok': True})


# ============================================================
# ROTAS — EXPORTAÇÃO
# ============================================================

from openpyxl.styles import (Font as OxFont, PatternFill as OxFill,
                              Alignment as OxAlign, Border as OxBorder, Side as OxSide)
from openpyxl.utils import get_column_letter as gcl

def _xl_fill(hex_color):   return OxFill("solid", fgColor=hex_color)
def _xl_font(bold=False, color="000000", size=10): return OxFont(name='Arial', bold=bold, color=color, size=size)
def _xl_border():
    s = OxSide(style='thin', color='CCCCCC')
    return OxBorder(left=s, right=s, top=s, bottom=s)
def _xl_align(h='left', wrap=False): return OxAlign(horizontal=h, vertical='center', wrap_text=wrap)

NAVY = "0D1838"; DGRAY = "DADADA"; WHITE = "FFFFFF"; BLACK = "000000"
LGRY = "F5F5F5"; LBLU = "D6E4F0"

DICT_COLUNAS = [
    ("Data","ORIGEM","Data do lançamento conforme o extrato original.","Formato YYYY-MM-DD. Ausente → Grupo B."),
    ("Valor","ORIGEM","Valor absoluto do lançamento (sempre positivo). Direção está em Tipo.","Numérico positivo. Ausente → Grupo B."),
    ("Tipo","ORIGEM","Direção do fluxo inferida pelo sistema.","entrada | saida | indefinido"),
    ("Descricao","ORIGEM","Histórico ou memo do lançamento conforme o banco.","Texto livre. Ausente → Grupo B."),
    ("Conta","ORIGEM","Número da conta bancária de origem, quando disponível.","Texto. Pode estar vazio."),
    ("Banco","ORIGEM","Nome da instituição financeira, quando disponível.","Texto livre."),
    ("CNPJ","ORIGEM","CNPJ ou CPF associado ao lançamento, quando disponível.","Texto. Pode conter CPF mascarado."),
    ("Centro_Custo","ORIGEM","Centro de custo, quando disponível ou atribuído manualmente.","Texto livre."),
    ("Marca","PROCESSAMENTO","Marca do grupo à qual o lançamento pertence.","Manta | RaiaClube | Ludika | GN | N/D"),
    ("Unidade","PROCESSAMENTO","Unidade de negócio específica identificada para o arquivo.","Ex: Manta 712, RaiaClube 608. N/D se não identificado."),
    ("Arquivo","PROCESSAMENTO","Nome do arquivo de extrato de origem.","Nome do .csv ou .xlsx enviado."),
    ("Grupo","PROCESSAMENTO","A = todos os campos ok e unidade identificada. B = revisão humana.","A | B"),
    ("Confiabilidade","PROCESSAMENTO","Grau de confiança do processamento automático.","ALTA | MÉDIA | BAIXA"),
    ("Status","PROCESSAMENTO","Estado final após o ciclo de validação.","confirmado | pendente | excluido"),
    ("Issues","PROCESSAMENTO","Motivos que levaram ao Grupo B. Vazio para Grupo A.","data_ausente | valor_ausente | descricao_ausente | tipo_indefinido | unidade_incerta | data_invalida | conflito_cabecalho_nome_arquivo | conta_nao_identificada"),
    ("ID","PROCESSAMENTO","Identificador único gerado pelo sistema. Formato: arquivo::linha.","Permite rastrear a linha exata no arquivo original."),
    ("Conta_Id","PROCESSAMENTO","ID da conta bancária resolvida a partir do cabeçalho do arquivo.","Referência para tabela-mãe de contas. Vazio em modo_legado."),
    ("Metodo_Atribuicao","PROCESSAMENTO","Como a conta/unidade foi atribuída (para auditoria).","cabecalho+arquivo_concordam | cabecalho_sem_confirmacao_nome | conflito_resolvido_manual | apenas_nome_confirmado_manual | modo_legado"),
    ("Confiab_Rastreabilidade","PROCESSAMENTO","Grau de confiança da atribuição da CONTA (distinta de Confiabilidade do lançamento).","ALTA: cabeçalho identificou sem conflito. MEDIA: resolvido manualmente pelo operador. BAIXA: modo_legado ou não identificado."),
]

COL_WIDTHS = {
    'Data':18,'Data_Raw':18,'Valor':14,'Tipo':12,'Descricao':38,'Conta':16,'Banco':28,
    'CNPJ':20,'Centro_Custo':18,'Marca':14,'Unidade':20,'Arquivo':38,
    'Grupo':10,'Confiabilidade':16,'Status':14,'Issues':40,'ID':42,
    'Conta_Id':18,'Metodo_Atribuicao':34,'Confiab_Rastreabilidade':14,
}


def _build_excel(ls: list, depara_cfg: dict, sumario_rows: list) -> io.BytesIO:
    """
    Constrói o Excel de exportação com 4 abas formatadas.

    Abas:
      Legenda     — Contexto, código de cores e dicionário de colunas.
      BD_Extratos — Dados normalizados. Cabeçalho duplo: banner + nomes.
                    Cinza = origem, azul escuro = processamento.
      De_Para     — Mapeamento coluna original → anchor aprendido.
      Sumario     — Métricas e totais financeiros da sessão.

    Returns: BytesIO posicionado no início, pronto para send_file().
    """
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    # Só expõe Data_Raw se houver data_invalida em algum lançamento —
    # caso contrário a coluna fica sempre vazia e só polui a planilha.
    has_data_invalida = any('data_invalida' in l.get('issues', []) for l in ls)
    COLS_ORIGEM_BASE = (
        ['Data','Data_Raw','Valor','Tipo','Descricao','Conta','Banco','CNPJ','Centro_Custo']
        if has_data_invalida else
        ['Data','Valor','Tipo','Descricao','Conta','Banco','CNPJ','Centro_Custo']
    )
    COLS_PROC = ['Marca','Unidade','Arquivo','Grupo','Confiabilidade',
                 'Confiab_Rastreabilidade','Conta_Id','Metodo_Atribuicao',
                 'Status','Issues','ID']

    extra_keys_all = {}
    for l in ls:
        for k in l.get('extras', {}):
            if 'Unnamed' not in k and k != 'extra_Nosso Número':
                v = l['extras'][k]
                if v not in (None, '', 'nan'):
                    extra_keys_all[k] = extra_keys_all.get(k, 0) + 1
    extra_keys    = [k for k, cnt in extra_keys_all.items() if cnt > 0]
    extra_display = {k: k.replace('extra_', '') for k in extra_keys}

    cols_data = COLS_ORIGEM_BASE + extra_keys + COLS_PROC
    cols_disp = COLS_ORIGEM_BASE + [extra_display[k] for k in extra_keys] + COLS_PROC
    n_orig = len(COLS_ORIGEM_BASE) + len(extra_keys)
    n_proc = len(COLS_PROC)
    ts_now = datetime.now().strftime('%d/%m/%Y %H:%M')

    # ── Aba Legenda ──────────────────────────────────────────────────────
    ws_leg = wb.create_sheet("Legenda")
    ws_leg.sheet_view.showGridLines = False

    def mw(ws, s, e, val, fnt, fll=None, aln=None, h=None):
        ws.merge_cells(f"{s}:{e}")
        c = ws[s]; c.value = val; c.font = fnt
        if fll: c.fill = fll
        if aln: c.alignment = aln
        r = int(''.join(filter(str.isdigit, s)))
        if h: ws.row_dimensions[r].height = h

    mw(ws_leg,"A1","D1","Guia de Leitura — BD_Extratos",
       _xl_font(bold=True,size=14,color=NAVY),_xl_fill("F0F4FA"),_xl_align(h='left'),28)
    mw(ws_leg,"A2","D2",f"Validação de Extratos v8.0 — Stoic Capital  |  {ts_now}",
       _xl_font(size=9,color="888888"),h=16)
    mw(ws_leg,"A4","D4","O QUE É ESTE ARQUIVO",
       _xl_font(bold=True,size=10,color=WHITE),_xl_fill(NAVY),_xl_align(h='left'))
    for r, txt in [
        (5,"Esta planilha consolida extratos bancários de diferentes unidades e marcas. "
           "Cada linha representa um lançamento bancário individual normalizado."),
        (6,"Colunas em dois blocos: ORIGEM (cinza) = dados dos extratos; "
           "PROCESSAMENTO (azul escuro) = atribuídos pelo sistema."),
        (7,"GRUPO A: validação automática completa. "
           "GRUPO B: inconsistência detectada, submetido a revisão humana."),
    ]:
        ws_leg.merge_cells(f"A{r}:D{r}")
        c = ws_leg[f"A{r}"]
        c.value = txt; c.font = _xl_font(size=9); c.fill = _xl_fill("F8F9FB")
        c.alignment = _xl_align(h='left', wrap=True); ws_leg.row_dimensions[r].height = 38

    mw(ws_leg,"A9","D9","CÓDIGO DE CORES",
       _xl_font(bold=True,size=10,color=WHITE),_xl_fill(NAVY),_xl_align(h='left'))
    for r,bg,fg,lbl,desc in [
        (10,DGRAY,BLACK,"ORIGEM","Dado proveniente diretamente do arquivo de extrato."),
        (11,NAVY,WHITE,"PROCESSAMENTO","Informação atribuída ou calculada pelo sistema."),
    ]:
        ws_leg.row_dimensions[r].height = 22
        c = ws_leg[f"A{r}"]
        c.value=lbl; c.font=_xl_font(bold=True,size=9,color=fg); c.fill=_xl_fill(bg); c.alignment=_xl_align(h='center')
        ws_leg.merge_cells(f"B{r}:D{r}")
        cd = ws_leg[f"B{r}"]; cd.value=desc; cd.font=_xl_font(size=9); cd.fill=_xl_fill("FAFAFA")

    mw(ws_leg,"A13","D13","DICIONÁRIO DE COLUNAS",
       _xl_font(bold=True,size=10,color=WHITE),_xl_fill(NAVY),_xl_align(h='left'))
    for ci,h in enumerate(["Coluna","Bloco","Descrição","Valores possíveis"],1):
        c = ws_leg.cell(row=14,column=ci)
        c.value=h; c.font=_xl_font(bold=True,size=9); c.fill=_xl_fill("E8E8E8"); c.alignment=_xl_align(h='center')
    for i,(col,bloco,desc,vals) in enumerate(DICT_COLUNAS):
        r=15+i; bg=LGRY if i%2==0 else WHITE; ws_leg.row_dimensions[r].height=30
        ca=ws_leg.cell(row=r,column=1)
        ca.value=col; ca.font=_xl_font(bold=True,size=9,color=WHITE if bloco=="PROCESSAMENTO" else BLACK)
        ca.fill=_xl_fill(NAVY if bloco=="PROCESSAMENTO" else DGRAY); ca.alignment=_xl_align()
        for ci2,v2 in enumerate([bloco,desc,vals],2):
            c2=ws_leg.cell(row=r,column=ci2)
            c2.value=v2; c2.font=_xl_font(size=9); c2.fill=_xl_fill(bg)
            c2.alignment=_xl_align(h='left' if ci2>2 else 'center',wrap=True)
    ws_leg.column_dimensions['A'].width=24; ws_leg.column_dimensions['B'].width=18
    ws_leg.column_dimensions['C'].width=55; ws_leg.column_dimensions['D'].width=52

    # ── Aba BD_Extratos ──────────────────────────────────────────────────
    ws = wb.create_sheet("BD_Extratos")
    ws.sheet_view.showGridLines = False

    # Linha 1: banner agrupador
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=n_orig)
    c1=ws.cell(row=1,column=1)
    c1.value="⬇  DADOS DE ORIGEM  —  provenientes dos arquivos de extrato"
    c1.font=_xl_font(bold=True,size=9,color=BLACK); c1.fill=_xl_fill(DGRAY); c1.alignment=_xl_align(h='center')
    ws.merge_cells(start_row=1,start_column=n_orig+1,end_row=1,end_column=n_orig+n_proc)
    c2=ws.cell(row=1,column=n_orig+1)
    c2.value="⬇  DADOS DE PROCESSAMENTO  —  atribuídos pelo sistema"
    c2.font=_xl_font(bold=True,size=9,color=WHITE); c2.fill=_xl_fill(NAVY); c2.alignment=_xl_align(h='center')
    ws.row_dimensions[1].height=20

    # Linha 2: cabeçalhos individuais
    for ci,(cd,dd) in enumerate(zip(cols_data,cols_disp),1):
        c=ws.cell(row=2,column=ci); is_p=cd in COLS_PROC
        c.value=dd; c.font=_xl_font(bold=True,size=9,color=WHITE if is_p else BLACK)
        c.fill=_xl_fill(NAVY if is_p else DGRAY); c.alignment=_xl_align(h='center'); c.border=_xl_border()
    ws.row_dimensions[2].height=22

    # Dados
    KEY_MAP = {'Marca':'marca','Unidade':'unidade','Arquivo':'arquivo',
               'Grupo':'grupo','Confiabilidade':'confiabilidade','Status':'status','ID':'id',
               'Conta_Id':'conta_id','Metodo_Atribuicao':'metodo_atribuicao',
               'Confiab_Rastreabilidade':'Confiab_Rastreabilidade'}
    # v9: índice por id para derivar a coluna 'Conta' a partir de contas_bancarias.
    contas_idx = {c['id']: c for c in CFG.get('contas_bancarias', [])}

    def _val_conta_origem(lanc):
        cid = lanc.get('conta_id')
        if cid and cid in contas_idx:
            c = contas_idx[cid]
            return c.get('conta') or lanc.get('Conta', '')
        return lanc.get('Conta', '')

    for ri,l in enumerate(ls,start=3):
        for ci,cd in enumerate(cols_data,1):
            is_p=cd in COLS_PROC
            if cd in COLS_PROC:
                val=(', '.join(l.get('issues',[])) if cd=='Issues' else l.get(KEY_MAP.get(cd,cd),''))
            elif cd in extra_keys:
                val=l.get('extras',{}).get(cd)
            elif cd == 'Conta':
                val = _val_conta_origem(l)
            else:
                val=l.get(cd)
            c=ws.cell(row=ri,column=ci)
            c.value=None if val in (None,'nan') else val
            c.font=_xl_font(size=9); c.alignment=_xl_align(); c.border=_xl_border()
            if ri%2==0: c.fill=_xl_fill(LBLU if is_p else LGRY)
            if cd=='Status':
                cm={'confirmado':"155724",'pendente':"78350F",'excluido':"888888"}
                c.font=_xl_font(size=9,color=cm.get(str(val),BLACK),bold=(str(val)=='confirmado'))
            if cd=='Grupo':
                c.font=_xl_font(size=9,bold=True,color="155724" if str(val)=='A' else "7F1D1D")
            if cd=='Tipo':
                c.font=_xl_font(size=9,color="1E3A5F" if str(val)=='entrada' else "7F1D1D" if str(val)=='saida' else BLACK)
    for ci,cd in enumerate(cols_data,1):
        nm=extra_display.get(cd,cd)
        ws.column_dimensions[gcl(ci)].width=COL_WIDTHS.get(cd,COL_WIDTHS.get(nm,18))
    ws.freeze_panes='A3'

    # ── Aba De_Para ──────────────────────────────────────────────────────
    ws_dp = wb.create_sheet("De_Para")
    ws_dp.sheet_view.showGridLines = False
    ws_dp.merge_cells("A1:C1"); ws_dp["A1"].value="Mapeamento de colunas — De-Para"
    ws_dp["A1"].font=_xl_font(bold=True,size=11,color=NAVY)
    ws_dp.merge_cells("A2:C2")
    ws_dp["A2"].value="Registra como cada coluna dos arquivos foi mapeada para o schema padronizado."
    ws_dp["A2"].font=_xl_font(size=9,color="666666"); ws_dp["A2"].alignment=_xl_align(h='left',wrap=True)
    ws_dp.row_dimensions[2].height=28
    for ci,h in enumerate(["Coluna Original","Anchor Normalizado","Origem"],1):
        c=ws_dp.cell(row=4,column=ci); c.value=h
        c.font=_xl_font(bold=True,size=9,color=WHITE); c.fill=_xl_fill(NAVY); c.alignment=_xl_align(h='center')
    for ri,(co,anc) in enumerate(depara_cfg.items(),start=5):
        bg=LGRY if ri%2==0 else WHITE
        for ci,v in enumerate([co,anc,'auto'],1):
            c=ws_dp.cell(row=ri,column=ci); c.value=v; c.font=_xl_font(size=9); c.fill=_xl_fill(bg)
    ws_dp.column_dimensions['A'].width=35; ws_dp.column_dimensions['B'].width=28; ws_dp.column_dimensions['C'].width=20

    # ── Aba Sumário ──────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Sumario")
    ws_sum.sheet_view.showGridLines = False
    ws_sum.merge_cells("A1:B1"); ws_sum["A1"].value="Sumário do processamento"
    ws_sum["A1"].font=_xl_font(bold=True,size=13,color=NAVY); ws_sum.row_dimensions[1].height=26
    ws_sum.merge_cells("A2:B2"); ws_sum["A2"].value=f"Gerado em: {ts_now}"
    ws_sum["A2"].font=_xl_font(size=9,color="888888")
    for ci,h in enumerate(["Métrica","Valor"],1):
        c=ws_sum.cell(row=4,column=ci); c.value=h
        c.font=_xl_font(bold=True,size=9,color=WHITE); c.fill=_xl_fill(NAVY); c.alignment=_xl_align(h='center')
    for ri,(m,v) in enumerate(sumario_rows,start=5):
        bg=LGRY if ri%2==0 else WHITE
        cm=ws_sum.cell(row=ri,column=1); cm.value=m; cm.font=_xl_font(size=10); cm.fill=_xl_fill(bg)
        cv=ws_sum.cell(row=ri,column=2); cv.value=v; cv.font=_xl_font(size=10,bold=True)
        cv.fill=_xl_fill(bg); cv.alignment=_xl_align(h='right')
    ws_sum.column_dimensions['A'].width=36; ws_sum.column_dimensions['B'].width=24

    wb.active = wb["Legenda"]
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out


@app.route('/api/exportar')
def exportar():
    """Gera e retorna o Excel formatado. Exclui lançamentos com status 'excluido'."""
    ls = [l for l in SESSION['lancamentos'] if l['status'] != 'excluido']
    if not ls:
        return jsonify({'ok': False, 'msg': 'Nenhum lançamento para exportar'}), 400
    sume = round(sum(l['Valor'] or 0 for l in ls if l['Tipo'] == 'entrada'), 2)
    sums = round(sum(l['Valor'] or 0 for l in ls if l['Tipo'] == 'saida'),   2)
    all_ls = SESSION['lancamentos']
    sumario_rows = [
        ('Total lançamentos exportados', len(ls)),
        ('Total na sessão',              len(all_ls)),
        ('Confirmados',  sum(1 for l in all_ls if l['status'] == 'confirmado')),
        ('Excluídos',    sum(1 for l in all_ls if l['status'] == 'excluido')),
        ('Pendentes',    sum(1 for l in all_ls if l['status'] == 'pendente')),
        ('Grupo A',      sum(1 for l in all_ls if l['grupo']  == 'A')),
        ('Grupo B',      sum(1 for l in all_ls if l['grupo']  == 'B')),
        ('Total Entradas (R$)', sume),
        ('Total Saídas (R$)',   sums),
        ('Saldo (R$)',          round(sume - sums, 2)),
        ('Gerado em',           datetime.now().strftime('%d/%m/%Y %H:%M')),
    ]
    out = _build_excel(ls, CFG.get('depara', {}), sumario_rows)
    ts  = datetime.now().strftime('%Y%m%d_%H%M')
    return send_file(out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=f'BD_Extratos_{ts}.xlsx')


# ============================================================
# ROTAS — CONTAS BANCÁRIAS (CRUD v9)
# ============================================================

_CAMPOS_CONTA_OBRIGATORIOS  = ('id', 'banco_nome', 'unit_id')
_CAMPOS_CONTA_OPCIONAIS     = ('banco_codigo', 'agencia', 'conta', 'cnpj_titular',
                               'ativo_desde', 'ativo_ate', 'observacao')


def _validar_conta_payload(d: dict, criar: bool = True) -> tuple[bool, str | None]:
    if criar:
        for k in _CAMPOS_CONTA_OBRIGATORIOS:
            if not d.get(k):
                return False, f'Campo obrigatório ausente: {k}'
    if d.get('unit_id') and not any(u['id'] == d['unit_id'] for u in CFG['unidades']):
        return False, f"unit_id '{d['unit_id']}' não existe em unidades"
    return True, None


@app.route('/api/contas', methods=['GET', 'POST', 'PUT', 'DELETE'])
def contas_bancarias():
    """CRUD da tabela-mãe de contas bancárias. Persiste em config.json."""
    if request.method == 'GET':
        return jsonify({'contas': CFG.get('contas_bancarias', [])})

    d = request.get_json() or {}

    if request.method == 'POST':
        ok, msg = _validar_conta_payload(d, criar=True)
        if not ok:
            return jsonify({'ok': False, 'msg': msg}), 400
        if any(c['id'] == d['id'] for c in CFG['contas_bancarias']):
            return jsonify({'ok': False, 'msg': 'ID já existe'}), 400
        nova = {
            'id':            d['id'],
            'banco_nome':    d['banco_nome'],
            'banco_codigo':  d.get('banco_codigo'),
            'agencia':       d.get('agencia'),
            'conta':         d.get('conta'),
            'cnpj_titular':  d.get('cnpj_titular'),
            'unit_id':       d['unit_id'],
            'ativo_desde':   d.get('ativo_desde'),
            'ativo_ate':     d.get('ativo_ate'),
            'observacao':    d.get('observacao', ''),
        }
        CFG['contas_bancarias'].append(nova)
        save_config(CFG)
        return jsonify({'ok': True, 'contas': CFG['contas_bancarias']})

    if request.method == 'PUT':
        if not d.get('id'):
            return jsonify({'ok': False, 'msg': 'id obrigatório'}), 400
        ok, msg = _validar_conta_payload(d, criar=False)
        if not ok:
            return jsonify({'ok': False, 'msg': msg}), 400
        for c in CFG['contas_bancarias']:
            if c['id'] == d['id']:
                for k in _CAMPOS_CONTA_OPCIONAIS + ('banco_nome', 'unit_id'):
                    if k in d:
                        c[k] = d[k]
                save_config(CFG)
                return jsonify({'ok': True, 'contas': CFG['contas_bancarias']})
        return jsonify({'ok': False, 'msg': 'Conta não encontrada'}), 404

    if request.method == 'DELETE':
        before = len(CFG['contas_bancarias'])
        CFG['contas_bancarias'] = [c for c in CFG['contas_bancarias'] if c['id'] != d.get('id')]
        if len(CFG['contas_bancarias']) < before:
            save_config(CFG)
            return jsonify({'ok': True})
        return jsonify({'ok': False, 'msg': 'Conta não encontrada'}), 404


@app.route('/api/confirmar-conta', methods=['POST'])
def confirmar_conta():
    """Resolve manualmente a atribuição conta↔arquivo.

    Body: { filename, conta_id, motivo }
    - conta_id=None cancela a confirmação prévia.
    - motivo é obrigatório quando o cruzamento atual é 'conflito' (escolha
      que diverge do cabeçalho). Guardado em conta_por_arquivo_meta.
    """
    d = request.get_json() or {}
    fn        = d.get('filename')
    conta_id  = d.get('conta_id')
    motivo    = (d.get('motivo') or '').strip()

    if not fn:
        return jsonify({'ok': False, 'msg': 'filename obrigatório'}), 400

    if conta_id is None or conta_id == '':
        SESSION['conta_por_arquivo'].pop(fn, None)
        SESSION['conta_por_arquivo_meta'].pop(fn, None)
        return jsonify({'ok': True, 'clearedfor': fn})

    conta = next((c for c in CFG['contas_bancarias'] if c['id'] == conta_id), None)
    if not conta:
        return jsonify({'ok': False, 'msg': f"conta_id '{conta_id}' não cadastrada"}), 404

    cruz = SESSION.get('cruzamento', {}).get(fn, {})
    if cruz.get('status') == 'conflito' and not motivo:
        return jsonify({
            'ok': False,
            'msg': 'Conflito cabeçalho×nome exige motivo por auditoria.'
        }), 400

    SESSION['conta_por_arquivo'][fn] = conta_id
    SESSION['conta_por_arquivo_meta'][fn] = {
        'metodo_original': cruz.get('metodo'),
        'escolha_humana':  conta_id,
        'motivo':          motivo,
        'timestamp':       datetime.now().isoformat(timespec='seconds'),
    }
    return jsonify({'ok': True, 'conta_id': conta_id})


# ============================================================
# ROTAS — UNIDADES (CRUD)
# ============================================================

@app.route('/api/unidades', methods=['GET','POST','PUT','DELETE'])
def unidades():
    """CRUD de unidades de negócio. Todas as mutações persistem em config.json."""
    if request.method == 'GET':
        return jsonify({'unidades': CFG['unidades']})
    d = request.get_json()
    if request.method == 'POST':
        if not all([d.get('id'), d.get('marca'), d.get('desc_unidade')]):
            return jsonify({'ok': False, 'msg': 'ID, Marca e Descrição são obrigatórios'}), 400
        if any(u['id'] == d['id'] for u in CFG['unidades']):
            return jsonify({'ok': False, 'msg': 'ID já existe'}), 400
        CFG['unidades'].append({'id': d['id'], 'marca': d['marca'], 'desc_unidade': d['desc_unidade']})
        save_config(CFG); return jsonify({'ok': True, 'unidades': CFG['unidades']})
    if request.method == 'PUT':
        for u in CFG['unidades']:
            if u['id'] == d.get('id'):
                u['marca'] = d.get('marca', u['marca'])
                u['desc_unidade'] = d.get('desc_unidade', u['desc_unidade'])
                save_config(CFG); return jsonify({'ok': True})
        return jsonify({'ok': False, 'msg': 'Unidade não encontrada'}), 404
    if request.method == 'DELETE':
        before = len(CFG['unidades'])
        CFG['unidades'] = [u for u in CFG['unidades'] if u['id'] != d.get('id')]
        if len(CFG['unidades']) < before:
            save_config(CFG); return jsonify({'ok': True})
        return jsonify({'ok': False, 'msg': 'Unidade não encontrada'}), 404


# ============================================================
# ROTAS — DE-PARA (CRUD)
# ============================================================

@app.route('/api/depara', methods=['GET','POST','PUT','DELETE'])
def gerenciar_depara():
    """CRUD do mapeamento de colunas. Todas as mutações persistem em config.json."""
    if request.method == 'GET':
        return jsonify({'depara': CFG.get('depara', {})})
    d = request.get_json()
    if request.method == 'POST':
        col, anchor = d.get('col_original','').strip(), d.get('anchor','').strip()
        if not col or not anchor:
            return jsonify({'ok': False, 'msg': 'col_original e anchor obrigatórios'}), 400
        if anchor not in ANCHOR_MAPS:
            return jsonify({'ok': False, 'msg': f'Anchor "{anchor}" não existe'}), 400
        CFG.setdefault('depara', {})[col] = anchor
        save_config(CFG); return jsonify({'ok': True, 'depara': CFG['depara']})
    if request.method == 'PUT':
        col, anchor = d.get('col_original','').strip(), d.get('anchor','').strip()
        if col not in CFG.get('depara', {}):
            return jsonify({'ok': False, 'msg': 'Entrada não encontrada'}), 404
        if anchor not in ANCHOR_MAPS:
            return jsonify({'ok': False, 'msg': f'Anchor "{anchor}" não existe'}), 400
        CFG['depara'][col] = anchor
        save_config(CFG); return jsonify({'ok': True, 'depara': CFG['depara']})
    if request.method == 'DELETE':
        col = d.get('col_original','').strip()
        if col in CFG.get('depara', {}):
            del CFG['depara'][col]; save_config(CFG)
        return jsonify({'ok': True, 'depara': CFG.get('depara', {})})

@app.route('/api/depara/reset', methods=['POST'])
def reset_depara():
    """Apaga todo o de-para aprendido. Redescobre na próxima execução."""
    CFG['depara'] = {}
    save_config(CFG)
    return jsonify({'ok': True})


# ============================================================
# ROTAS — RESET DE SESSÃO
# ============================================================

@app.route('/api/limpar', methods=['POST'])
def limpar():
    """Reinicia sessão e remove arquivos do disco. Preserva configurações."""
    global SESSION
    SESSION = {
        'arquivos': [], 'lancamentos': [], 'schema_map': {},
        'processado': False, 'doc_verificados': {}, 'previews': {},
        'progresso': {'pct': 0, 'msg': '', 'ativo': False},
        'deteccao_cab': {}, 'cruzamento': {},
        'conta_por_arquivo': {}, 'conta_por_arquivo_meta': {},
        'arquivos_bloqueados': [], 'audit_log': [],
    }
    for f in UPLOAD_FOLDER.glob('*'):
        try: f.unlink()
        except: pass
    return jsonify({'ok': True})


# ============================================================
# ENTRY POINT
# ============================================================

if __name__ == '__main__':
    print("\n" + "=" * 60)
    print("  VALIDAÇÃO DE EXTRATOS v8.0 — Stoic Capital")
    print("  → http://127.0.0.1:5000")
    print("=" * 60 + "\n")
    app.run(debug=True, host='127.0.0.1', port=5000, use_reloader=False, threaded=True)
